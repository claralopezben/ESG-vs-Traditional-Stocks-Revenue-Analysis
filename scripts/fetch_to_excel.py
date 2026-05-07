"""
fetch_to_excel.py

Downloads ETF prices and Fama-French factors, calculates returns and
summary statistics, and saves everything into a single Excel workbook
with one sheet per data type.
"""
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

ESG_TICKERS = {
    "ESGU": "iShares ESG Aware MSCI USA ETF",
    "SUSA": "iShares MSCI USA ESG Select ETF",
    "ICLN": "iShares Global Clean Energy ETF",
}

TRADITIONAL_TICKERS = {
    "SPY":  "SPDR S&P 500 ETF Trust",
    "VTI":  "Vanguard Total Stock Market ETF",
    "DIA":  "SPDR Dow Jones Industrial Average ETF",
}

ALL_TICKERS  = {**ESG_TICKERS, **TRADITIONAL_TICKERS}
TRADING_DAYS = 252


def fetch_prices(tickers: list[str], start: str, end: str) -> pd.DataFrame:
    import yfinance as yf
    print(f"[1/3] Downloading {len(tickers)} tickers from Yahoo Finance ...")
    raw = yf.download(
        tickers=tickers,
        start=start,
        end=end,
        auto_adjust=True,
        progress=False,
        group_by="ticker",
    )
    if isinstance(raw.columns, pd.MultiIndex):
        prices = pd.DataFrame({t: raw[t]["Close"] for t in tickers})
    else:
        prices = raw[["Close"]].rename(columns={"Close": tickers[0]})
    prices.index.name = "Date"
    return prices.sort_index()


def fetch_fama_french(start: str, end: str) -> pd.DataFrame:
    from pandas_datareader import data as pdr
    print("[2/3] Downloading Fama-French 3 factors ...")
    ff = pdr.DataReader(
        "F-F_Research_Data_Factors_daily",
        "famafrench",
        start=start,
        end=end,
    )[0] / 100.0
    ff.index = pd.to_datetime(ff.index)
    ff.index.name = "Date"
    return ff


def build_dataset(prices: pd.DataFrame,
                  factors: pd.DataFrame) -> dict[str, pd.DataFrame]:
    print("[3/3] Aligning, computing returns and excess returns ...")

    common_idx = prices.index.intersection(factors.index)
    prices  = prices.loc[common_idx].dropna(how="any")
    factors = factors.loc[prices.index]

    simple_ret = prices.pct_change().dropna(how="any")
    log_ret    = np.log(prices / prices.shift(1)).dropna(how="any")

    rf = factors["RF"].reindex(simple_ret.index)
    excess_ret = simple_ret.sub(rf, axis=0)

    rows = []
    for tkr in simple_ret.columns:
        r  = simple_ret[tkr]
        er = excess_ret[tkr]
        ann_mean = r.mean()  * TRADING_DAYS
        ann_vol  = r.std(ddof=1) * np.sqrt(TRADING_DAYS)
        sharpe   = (er.mean() / er.std(ddof=1)) * np.sqrt(TRADING_DAYS)
        rows.append({
            "Ticker":           tkr,
            "Name":             ALL_TICKERS[tkr],
            "Group":            "ESG" if tkr in ESG_TICKERS else "Traditional",
            "N_obs":            int(len(r)),
            "Mean_Daily":       r.mean(),
            "Std_Daily":        r.std(ddof=1),
            "Ann_Return":       ann_mean,
            "Ann_Volatility":   ann_vol,
            "Sharpe_Ratio":     sharpe,
            "Min_Daily":        r.min(),
            "Max_Daily":        r.max(),
            "Skewness":         r.skew(),
            "Kurtosis":         r.kurtosis(),
        })
    summary = pd.DataFrame(rows).set_index("Ticker")

    tickers_table = pd.DataFrame(
        [(t, n, "ESG" if t in ESG_TICKERS else "Traditional")
         for t, n in ALL_TICKERS.items()],
        columns=["Ticker", "Name", "Group"],
    ).set_index("Ticker")

    return {
        "Prices":             prices,
        "Returns_Simple":     simple_ret,
        "Returns_Log":        log_ret,
        "Factors":            factors,
        "Excess_Returns":     excess_ret,
        "Summary_Statistics": summary,
        "Tickers":            tickers_table,
    }


def make_readme(start: str, end: str, n_obs: int) -> pd.DataFrame:
    rows = [
        ("Project",            "BEE2041 - ESG vs Traditional ETFs"),
        ("Created",             datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Sample window",       f"{start}  to  {end}"),
        ("Trading days in set", str(n_obs)),
        ("Data source 1",       "Yahoo Finance (yfinance) - adjusted close prices"),
        ("Data source 2",       "Kenneth French Data Library - FF3 factors + RF"),
        ("",                    ""),
        ("Sheet: Prices",            "Daily adjusted-close prices for the 6 ETFs."),
        ("Sheet: Returns_Simple",    "Daily simple returns."),
        ("Sheet: Returns_Log",       "Daily log returns."),
        ("Sheet: Factors",           "Fama-French daily factors and risk-free rate."),
        ("Sheet: Excess_Returns",    "Daily return minus risk-free rate per ETF."),
        ("Sheet: Summary_Statistics","Annualised stats per ETF."),
        ("Sheet: Tickers",           "Ticker to fund name mapping."),
    ]
    return pd.DataFrame(rows, columns=["Field", "Description"])


def write_excel(sheets: dict[str, pd.DataFrame],
                readme: pd.DataFrame,
                out_path: Path) -> None:
    from openpyxl.styles import Font, Alignment

    print(f"\nWriting workbook -> {out_path}")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name)

        wb = writer.book
        bold = Font(bold=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = bold
                cell.alignment = Alignment(horizontal="left")
            ws.freeze_panes = "B2"
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = max(
                    (len(str(c.value)) for c in col_cells if c.value is not None),
                    default=10,
                )
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = min(max_len + 2, 60)

    print(f"Done. {len(sheets) + 1} sheets written.")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--start", default="2019-01-01")
    p.add_argument("--end",   default="2025-12-31")
    p.add_argument("--out",   default="esg_dataset.xlsx")
    args = p.parse_args()

    out_path = Path(args.out).resolve()
    prices  = fetch_prices(list(ALL_TICKERS), args.start, args.end)
    factors = fetch_fama_french(args.start, args.end)
    sheets  = build_dataset(prices, factors)
    readme  = make_readme(args.start, args.end, n_obs=len(sheets["Returns_Simple"]))
    write_excel(sheets, readme, out_path)

    print("\n----- quick sanity check -----")
    print(sheets["Summary_Statistics"][
        ["Group", "Ann_Return", "Ann_Volatility", "Sharpe_Ratio"]
    ].round(4))


if __name__ == "__main__":
    main()
